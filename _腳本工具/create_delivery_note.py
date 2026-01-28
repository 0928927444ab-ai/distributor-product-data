# -*- coding: utf-8 -*-
"""
建立樂莫集團出貨單 Excel 模板
"""

import subprocess
import sys

# 嘗試安裝 openpyxl
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def create_delivery_note():
    wb = Workbook()
    ws = wb.active
    ws.title = "出貨單"

    # 定義樣式
    # 藍色標題背景白字
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    white_font_small = Font(color="FFFFFF", bold=True, size=11)

    # 淺黃色背景（可輸入欄位）
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # 淺藍色背景（次標題）
    light_blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # 邊框樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 字體樣式
    title_font = Font(bold=True, size=20)
    subtitle_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    # 置中對齊
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # 設定欄寬
    column_widths = {
        'A': 6,   # 項次
        'B': 20,  # 產品名稱
        'C': 18,  # 型號/規格
        'D': 15,  # 顏色/花色
        'E': 8,   # 單位
        'F': 10,  # 數量
        'G': 12,  # 單價
        'H': 14,  # 金額
        'I': 10,  # 箱號
        'J': 18,  # 備註
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ========== 1. 表頭區 ==========
    # 公司名稱
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "樂莫集團 摩瑪建材"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].font = Font(bold=True, size=20, color="FFFFFF")
    ws.row_dimensions[row].height = 35
    row += 1

    # 文件標題
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "出貨單 DELIVERY NOTE"
    ws[f'A{row}'].font = Font(bold=True, size=16)
    ws[f'A{row}'].alignment = center_align
    ws.row_dimensions[row].height = 28
    row += 1

    # 單據編號、出貨日期
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "單據編號："
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = right_align

    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'].fill = yellow_fill
    ws[f'C{row}'].border = thin_border
    ws[f'D{row}'].border = thin_border

    ws.merge_cells(f'G{row}:H{row}')
    ws[f'G{row}'] = "出貨日期："
    ws[f'G{row}'].font = header_font
    ws[f'G{row}'].alignment = right_align

    ws.merge_cells(f'I{row}:J{row}')
    ws[f'I{row}'].fill = yellow_fill
    ws[f'I{row}'].border = thin_border
    ws[f'J{row}'].border = thin_border

    ws.row_dimensions[row].height = 25
    row += 2

    # ========== 2. 客戶資訊區 ==========
    # 區塊標題
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "客戶資訊 CUSTOMER INFORMATION"
    ws[f'A{row}'].font = white_font_small
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 客戶資訊欄位
    customer_fields = [
        [("客戶名稱：", 2), (None, 3), ("聯絡電話：", 2), (None, 3)],
        [("收貨人：", 2), (None, 3), ("供應商：", 2), (None, 3)],
        [("送貨地址：", 2), (None, 8)],
        [("訂單編號：", 2), (None, 3), ("產品類別：", 2), (None, 3)],
    ]

    for fields in customer_fields:
        col = 1
        for field in fields:
            label, span = field
            start_col = get_column_letter(col)
            end_col = get_column_letter(col + span - 1)

            if span > 1:
                ws.merge_cells(f'{start_col}{row}:{end_col}{row}')

            if label:
                ws[f'{start_col}{row}'] = label
                ws[f'{start_col}{row}'].font = header_font
                ws[f'{start_col}{row}'].fill = light_blue_fill
                ws[f'{start_col}{row}'].alignment = right_align
            else:
                ws[f'{start_col}{row}'].fill = yellow_fill

            # 設定邊框
            for c in range(col, col + span):
                ws[f'{get_column_letter(c)}{row}'].border = thin_border

            col += span

        ws.row_dimensions[row].height = 25
        row += 1

    row += 1

    # ========== 3. 產品明細區 ==========
    # 區塊標題
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "產品明細 PRODUCT DETAILS"
    ws[f'A{row}'].font = white_font_small
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 表頭
    headers = ["項次", "產品名稱", "型號/規格", "顏色/花色", "單位", "數量", "單價", "金額", "箱號", "備註"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = light_blue_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 產品明細行（15行空白）
    detail_start_row = row
    for i in range(15):
        ws.cell(row=row, column=1).value = i + 1  # 項次
        for col_idx in range(1, 11):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_align
            if col_idx > 1:  # 除了項次外，其他欄位可輸入
                cell.fill = yellow_fill
        ws.row_dimensions[row].height = 22
        row += 1
    detail_end_row = row - 1

    # 合計行
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "合 計"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = center_align

    for col_idx in range(1, 11):
        ws.cell(row=row, column=col_idx).border = thin_border

    # 數量合計（F欄）
    ws.cell(row=row, column=6).value = f"=SUM(F{detail_start_row}:F{detail_end_row})"
    ws.cell(row=row, column=6).font = header_font
    ws.cell(row=row, column=6).alignment = center_align
    ws.cell(row=row, column=6).fill = light_blue_fill

    # 金額合計（H欄）
    ws.cell(row=row, column=8).value = f"=SUM(H{detail_start_row}:H{detail_end_row})"
    ws.cell(row=row, column=8).font = header_font
    ws.cell(row=row, column=8).alignment = center_align
    ws.cell(row=row, column=8).fill = light_blue_fill

    ws.row_dimensions[row].height = 25
    row += 2

    # ========== 4. 包裝資訊區 ==========
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "包裝資訊 PACKAGING INFORMATION"
    ws[f'A{row}'].font = white_font_small
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 包裝欄位
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "總箱數/托數："
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = right_align
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border

    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'].fill = yellow_fill
    ws[f'C{row}'].border = thin_border
    ws[f'D{row}'].border = thin_border

    ws[f'E{row}'] = "總體積(m³)："
    ws[f'E{row}'].font = header_font
    ws[f'E{row}'].fill = light_blue_fill
    ws[f'E{row}'].alignment = right_align
    ws[f'E{row}'].border = thin_border

    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'].fill = yellow_fill
    ws[f'F{row}'].border = thin_border
    ws[f'G{row}'].border = thin_border

    ws[f'H{row}'] = "總重量(kg)："
    ws[f'H{row}'].font = header_font
    ws[f'H{row}'].fill = light_blue_fill
    ws[f'H{row}'].alignment = right_align
    ws[f'H{row}'].border = thin_border

    ws.merge_cells(f'I{row}:J{row}')
    ws[f'I{row}'].fill = yellow_fill
    ws[f'I{row}'].border = thin_border
    ws[f'J{row}'].border = thin_border

    ws.row_dimensions[row].height = 25
    row += 2

    # ========== 5. 備註區 ==========
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "備 註 REMARKS"
    ws[f'A{row}'].font = white_font_small
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 備註輸入區（合併4行）
    ws.merge_cells(f'A{row}:J{row+3}')
    ws[f'A{row}'].fill = yellow_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for r in range(row, row + 4):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = thin_border
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row+1].height = 20
    ws.row_dimensions[row+2].height = 20
    ws.row_dimensions[row+3].height = 20
    row += 5

    # ========== 6. 簽收區 ==========
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "簽收確認 SIGNATURE CONFIRMATION"
    ws[f'A{row}'].font = white_font_small
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].border = thin_border
    ws.row_dimensions[row].height = 25
    row += 1

    # 簽收欄位
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "出貨人簽章："
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = right_align
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border

    ws.merge_cells(f'C{row}:E{row}')
    ws[f'C{row}'].fill = yellow_fill
    for c in ['C', 'D', 'E']:
        ws[f'{c}{row}'].border = thin_border

    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'] = "收貨人簽章："
    ws[f'F{row}'].font = header_font
    ws[f'F{row}'].fill = light_blue_fill
    ws[f'F{row}'].alignment = right_align
    ws[f'F{row}'].border = thin_border
    ws[f'G{row}'].border = thin_border

    ws.merge_cells(f'H{row}:J{row}')
    ws[f'H{row}'].fill = yellow_fill
    for c in ['H', 'I', 'J']:
        ws[f'{c}{row}'].border = thin_border

    ws.row_dimensions[row].height = 40
    row += 1

    # 日期欄位
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "出貨日期："
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = light_blue_fill
    ws[f'A{row}'].alignment = right_align
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].border = thin_border

    ws.merge_cells(f'C{row}:E{row}')
    ws[f'C{row}'].fill = yellow_fill
    for c in ['C', 'D', 'E']:
        ws[f'{c}{row}'].border = thin_border

    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'] = "簽收日期："
    ws[f'F{row}'].font = header_font
    ws[f'F{row}'].fill = light_blue_fill
    ws[f'F{row}'].alignment = right_align
    ws[f'F{row}'].border = thin_border
    ws[f'G{row}'].border = thin_border

    ws.merge_cells(f'H{row}:J{row}')
    ws[f'H{row}'].fill = yellow_fill
    for c in ['H', 'I', 'J']:
        ws[f'{c}{row}'].border = thin_border

    ws.row_dimensions[row].height = 25
    row += 2

    # ========== 7. 頁尾 ==========
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "樂莫集團 摩瑪建材 | 電話：(02) XXXX-XXXX | 傳真：(02) XXXX-XXXX | 地址：台北市XX區XX路XX號"
    ws[f'A{row}'].font = Font(size=9, color="666666")
    ws[f'A{row}'].alignment = center_align
    ws.row_dimensions[row].height = 20

    # 設定列印區域
    ws.print_area = f'A1:J{row}'

    # 設定頁面邊距
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # 儲存檔案
    output_dir = r"C:\Users\User\Desktop\總代理產品資料\00_確認落地產品"
    os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, "樂莫集團_摩瑪建材_通用出貨單.xlsx")
    wb.save(output_path)
    print(f"出貨單模板已成功建立！")
    print(f"檔案位置：{output_path}")
    return output_path

if __name__ == "__main__":
    create_delivery_note()

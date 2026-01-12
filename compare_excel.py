# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter

file1 = r'c:\Users\User\Desktop\總代理產品資料\04_五金\威斯特-VSTE\門用五金\03_訂單管理\20251114第二批庫存\摩玛(威思特产品）报价单2025-11-14(調整後)(2).xlsx'
file2 = r'c:\Users\User\Desktop\總代理產品資料\04_五金\威斯特-VSTE\門用五金\03_訂單管理\20251201下單資料\摩玛(威思特产品）报价单2025-11-14(調整後)(2)(1).xlsx'

wb1 = openpyxl.load_workbook(file1, data_only=True)
wb2 = openpyxl.load_workbook(file2, data_only=True)

print("=" * 80)
print("Excel File Comparison Result / Excel 檔案比較結果")
print("=" * 80)

sheets1 = wb1.sheetnames
sheets2 = wb2.sheetnames

print()
print("[1] Sheet Count and Names / 工作表數量和名稱")
print(f"File1 sheets: {len(sheets1)} - {sheets1}")
print(f"File2 sheets: {len(sheets2)} - {sheets2}")

sheets_same = sheets1 == sheets2
if sheets_same:
    print("[OK] Sheet names are identical / 工作表名稱完全相同")
else:
    print("[DIFF] Sheet names differ / 工作表名稱不同")
    only_in_1 = set(sheets1) - set(sheets2)
    only_in_2 = set(sheets2) - set(sheets1)
    if only_in_1:
        print(f"  Only in File1: {only_in_1}")
    if only_in_2:
        print(f"  Only in File2: {only_in_2}")

print()
print("[2] Row/Column Dimensions / 行列數比較")
print("[3] Cell-by-cell Comparison / 逐格比較")

common_sheets = [s for s in sheets1 if s in sheets2]
all_diffs = []
dim_same = True

for sn in common_sheets:
    ws1 = wb1[sn]
    ws2 = wb2[sn]
    print()
    print(f"--- Sheet: {sn} ---")
    r1, c1 = ws1.max_row, ws1.max_column
    r2, c2 = ws2.max_row, ws2.max_column
    print(f"File1: {r1} rows x {c1} cols")
    print(f"File2: {r2} rows x {c2} cols")

    if r1 == r2 and c1 == c2:
        print("[OK] Same dimensions / 行列數相同")
    else:
        dim_same = False
        print("[DIFF] Different dimensions / 行列數不同")

    mr = max(r1, r2)
    mc = max(c1, c2)
    diffs = []

    for row in range(1, mr + 1):
        for col in range(1, mc + 1):
            v1 = ws1.cell(row=row, column=col).value
            v2 = ws2.cell(row=row, column=col).value
            if v1 != v2:
                cell = f'{get_column_letter(col)}{row}'
                diffs.append((cell, v1, v2))

    if diffs:
        print(f"[DIFF] Found {len(diffs)} differences / 發現 {len(diffs)} 處差異:")
        for i, (cell, v1, v2) in enumerate(diffs[:50]):
            v1_str = str(v1) if v1 is not None else "(empty)"
            v2_str = str(v2) if v2 is not None else "(empty)"
            if len(v1_str) > 50:
                v1_str = v1_str[:50] + "..."
            if len(v2_str) > 50:
                v2_str = v2_str[:50] + "..."
            print(f"  {cell}: File1=\"{v1_str}\" vs File2=\"{v2_str}\"")
        if len(diffs) > 50:
            print(f"  ... and {len(diffs) - 50} more differences")
        all_diffs.extend([(sn, d) for d in diffs])
    else:
        print("[OK] Content identical / 內容完全相同")

print()
print("=" * 80)
print("[Summary / 總結]")
print("=" * 80)
print()

if sheets_same and dim_same and len(all_diffs) == 0:
    print("CONCLUSION / 結論:")
    print("The two Excel files are IDENTICAL!")
    print("兩個 Excel 檔案內容完全相同!")
else:
    print("CONCLUSION / 結論:")
    print("The two Excel files have DIFFERENCES")
    print("兩個 Excel 檔案存在差異")
    print()
    if not sheets_same:
        print("- Sheet structure differs / 工作表結構不同")
    if not dim_same:
        print("- Some sheets have different dimensions / 部分工作表行列數不同")
    if all_diffs:
        print(f"- Total cell differences: {len(all_diffs)} / 共 {len(all_diffs)} 處儲存格差異")

print()
wb1.close()
wb2.close()

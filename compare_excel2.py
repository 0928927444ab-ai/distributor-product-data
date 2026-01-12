# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import sys

# Force UTF-8 output
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file1 = r'c:\Users\User\Desktop\總代理產品資料\04_五金\威斯特-VSTE\門用五金\03_訂單管理\20251114第二批庫存\摩玛(威思特产品）报价单2025-11-14(調整後)(2).xlsx'
file2 = r'c:\Users\User\Desktop\總代理產品資料\04_五金\威斯特-VSTE\門用五金\03_訂單管理\20251201下單資料\摩玛(威思特产品）报价单2025-11-14(調整後)(2)(1).xlsx'

# Output to file
output_file = r'c:\Users\User\Desktop\總代理產品資料\comparison_result.txt'

with open(output_file, 'w', encoding='utf-8') as f:
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)

    f.write("=" * 80 + "\n")
    f.write("Excel File Comparison Result / Excel 檔案比較結果\n")
    f.write("=" * 80 + "\n")
    f.write("\n")
    f.write(f"File 1: {file1}\n")
    f.write(f"File 2: {file2}\n")
    f.write("\n")

    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    f.write("[1] Sheet Count and Names / 工作表數量和名稱\n")
    f.write(f"File1 sheets: {len(sheets1)} - {sheets1}\n")
    f.write(f"File2 sheets: {len(sheets2)} - {sheets2}\n")

    sheets_same = sheets1 == sheets2
    if sheets_same:
        f.write("[OK] Sheet names are identical / 工作表名稱完全相同\n")
    else:
        f.write("[DIFF] Sheet names differ / 工作表名稱不同\n")
        only_in_1 = set(sheets1) - set(sheets2)
        only_in_2 = set(sheets2) - set(sheets1)
        if only_in_1:
            f.write(f"  Only in File1: {only_in_1}\n")
        if only_in_2:
            f.write(f"  Only in File2: {only_in_2}\n")

    f.write("\n")
    f.write("[2] Row/Column Dimensions / 行列數比較\n")
    f.write("[3] Cell-by-cell Comparison / 逐格比較\n")

    common_sheets = [s for s in sheets1 if s in sheets2]
    all_diffs = []
    dim_same = True

    for sn in common_sheets:
        ws1 = wb1[sn]
        ws2 = wb2[sn]
        f.write("\n")
        f.write(f"--- Sheet: {sn} ---\n")
        r1, c1 = ws1.max_row, ws1.max_column
        r2, c2 = ws2.max_row, ws2.max_column
        f.write(f"File1: {r1} rows x {c1} cols\n")
        f.write(f"File2: {r2} rows x {c2} cols\n")

        if r1 == r2 and c1 == c2:
            f.write("[OK] Same dimensions / 行列數相同\n")
        else:
            dim_same = False
            f.write("[DIFF] Different dimensions / 行列數不同\n")

        mr = max(r1, r2)
        mc = max(c1, c2)
        diffs = []

        for row in range(1, mr + 1):
            for col in range(1, mc + 1):
                v1 = ws1.cell(row=row, column=col).value
                v2 = ws2.cell(row=row, column=col).value
                if v1 != v2:
                    cell = f'{get_column_letter(col)}{row}'
                    diffs.append((cell, v1, v2))

        if diffs:
            f.write(f"[DIFF] Found {len(diffs)} differences / 發現 {len(diffs)} 處差異:\n")
            for i, (cell, v1, v2) in enumerate(diffs[:50]):
                v1_str = str(v1) if v1 is not None else "(empty)"
                v2_str = str(v2) if v2 is not None else "(empty)"
                if len(v1_str) > 60:
                    v1_str = v1_str[:60] + "..."
                if len(v2_str) > 60:
                    v2_str = v2_str[:60] + "..."
                f.write(f"  {cell}: File1=\"{v1_str}\" vs File2=\"{v2_str}\"\n")
            if len(diffs) > 50:
                f.write(f"  ... and {len(diffs) - 50} more differences\n")
            all_diffs.extend([(sn, d) for d in diffs])
        else:
            f.write("[OK] Content identical / 內容完全相同\n")

    f.write("\n")
    f.write("=" * 80 + "\n")
    f.write("[Summary / 總結]\n")
    f.write("=" * 80 + "\n")
    f.write("\n")

    if sheets_same and dim_same and len(all_diffs) == 0:
        f.write("CONCLUSION / 結論:\n")
        f.write(">>> The two Excel files are IDENTICAL! <<<\n")
        f.write(">>> 兩個 Excel 檔案內容完全相同! <<<\n")
    else:
        f.write("CONCLUSION / 結論:\n")
        f.write(">>> The two Excel files have DIFFERENCES <<<\n")
        f.write(">>> 兩個 Excel 檔案存在差異 <<<\n")
        f.write("\n")
        if not sheets_same:
            f.write("- Sheet structure differs / 工作表結構不同\n")
        if not dim_same:
            f.write("- Some sheets have different dimensions / 部分工作表行列數不同\n")
        if all_diffs:
            f.write(f"- Total cell differences: {len(all_diffs)} / 共 {len(all_diffs)} 處儲存格差異\n")

    f.write("\n")
    wb1.close()
    wb2.close()

print(f"Result written to: {output_file}")
